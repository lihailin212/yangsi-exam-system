import json
import re
import io
import os
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from ..database import get_db
from .. import crud, schemas, models
from ..auth import get_current_user, require_admin

router = APIRouter(prefix="/api/questions", tags=["questions"])

@router.get("")
def list_questions(skip: int = 0, limit: int = 20, type: str = "", search: str = "", category: str = "",
                   db: Session = Depends(get_db), _=Depends(get_current_user)):
    items, total = crud.get_questions(db, skip=skip, limit=limit, q_type=type, search=search, category=category)
    return {"items": [schemas.QuestionOut.model_validate(q) for q in items], "total": total}

@router.post("", response_model=schemas.QuestionOut)
def create_question(q_in: schemas.QuestionCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    return crud.create_question(db, q_in)

@router.put("/{question_id}", response_model=schemas.QuestionOut)
def update_question(question_id: int, q_in: schemas.QuestionUpdate, db: Session = Depends(get_db), _=Depends(require_admin)):
    q = crud.update_question(db, question_id, q_in)
    if not q:
        raise HTTPException(status_code=404, detail="题目不存在")
    return q

@router.delete("/{question_id}")
def delete_question(question_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    crud.delete_question(db, question_id)
    return {"ok": True}

@router.post("/import")
async def import_questions(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_admin)):
    content = await file.read()
    filename = file.filename.lower()
    questions = []
    
    type_map_reverse = {"单选题": "single", "多选题": "multiple", "判断题": "judgment", "单选": "single", "多选": "multiple", "判断": "judgment"}

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # 跳过表头
            if not row[0]:
                continue
            
            q_type_raw = str(row[0] or "single").strip()
            q_type = type_map_reverse.get(q_type_raw, q_type_raw if q_type_raw in ("single", "multiple", "judgment") else "single")
            
            content_text = str(row[1] or "").strip()
            options_str = str(row[2] or "").strip()
            answer = str(row[3] or "").strip()
            analysis = str(row[4] or "").strip()
            # 注意：分值现在在考试层面设置，这里忽略导入的分值字段
            # score = int(row[5]) if row[5] and str(row[5]).isdigit() else 10
            category = str(row[6] or "").strip() if len(row) > 6 else ""
            img_paths = str(row[7] or "").strip() if len(row) > 7 else ""
            
            if img_paths:
                content_html = f"<p>{content_text}</p>"
                for img_path in img_paths.split("\n"):
                    img_path = img_path.strip()
                    if img_path:
                        content_html += f'<img src="{img_path}" alt="image" style="max-width: 50%;" />'
                content_text = content_html
            
            options = _parse_options(options_str)
            questions.append({
                "type": q_type,
                "content": content_text,
                "options": json.dumps(options, ensure_ascii=False),
                "answer": answer,
                "analysis": analysis,
                "category": category
            })

    elif filename.endswith(".txt") or filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = line.split(",", 6)
            if len(parts) < 4:
                continue
            q_type_raw = parts[0].strip()
            q_type = type_map_reverse.get(q_type_raw, q_type_raw if q_type_raw in ("single", "multiple", "judgment") else "single")
            content_text = parts[1].strip()
            options_str = parts[2].strip() if len(parts) > 2 else ""
            answer = parts[3].strip() if len(parts) > 3 else ""
            analysis = parts[4].strip() if len(parts) > 4 else ""
            # 注意：分值现在在考试层面设置，这里忽略导入的分值字段
            # score = int(parts[5].strip()) if len(parts) > 5 and parts[5].strip().isdigit() else 10
            category = parts[6].strip() if len(parts) > 6 else ""
            options = _parse_options(options_str)
            questions.append({"type": q_type, "content": content_text, "options": json.dumps(options, ensure_ascii=False),
                               "answer": answer, "analysis": analysis, "category": category})
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传 xlsx/xls/txt/csv")

    count = crud.bulk_create_questions(db, questions)
    return {"imported": count}

def _parse_options(options_str: str) -> dict:
    """解析 'A:选项1|B:选项2' 格式"""
    result = {}
    if not options_str:
        return result
    for part in options_str.split("|"):
        part = part.strip()
        if ":" in part:
            k, v = part.split(":", 1)
            result[k.strip()] = v.strip()
    return result

@router.get("/export")
async def export_questions(type: str = "", search: str = "", category: str = "",
                          db: Session = Depends(get_db), _=Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.drawing.image import Image as XLImage
    
    items, _ = crud.get_questions(db, skip=0, limit=10000, q_type=type, search=search, category=category)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目导出"
    
    headers = ["题型", "题目内容", "选项", "答案", "解析", "分类", "图片路径"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    type_map = {"single": "单选题", "multiple": "多选题", "judgment": "判断题"}
    
    for row_idx, q in enumerate(items, 2):
        q_type = type_map.get(q.type, q.type)
        content_html = q.content or ""
        
        img_urls = re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', content_html)
        img_display = "\n".join(img_urls) if img_urls else ""
        
        plain_text = re.sub(r'<[^>]+>', '', content_html).strip()
        
        try:
            options = json.loads(q.options) if q.options else {}
            options_str = "|".join([f"{k}:{v}" for k, v in options.items()])
        except:
            options_str = q.options or ""
        
        answer = q.answer or ""
        analysis = q.analysis or ""
        cat = q.category or ""

        ws.cell(row=row_idx, column=1, value=q_type)
        ws.cell(row=row_idx, column=2, value=plain_text)
        ws.cell(row=row_idx, column=3, value=options_str)
        ws.cell(row=row_idx, column=4, value=answer)
        ws.cell(row=row_idx, column=5, value=analysis)
        ws.cell(row=row_idx, column=6, value=cat)
        ws.cell(row=row_idx, column=7, value=img_display)

        img_row = row_idx + 1
        for img_url in img_urls:
            try:
                img_filename = os.path.basename(img_url)
                img_path = os.path.join(os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), 'uploads', 'images', img_filename)
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = 300
                    img.height = 200
                    ws.add_image(img, f'B{img_row}')
                    ws.row_dimensions[img_row].height = 130
                    break
            except Exception as e:
                print(f"Error: {e}")
                pass
    
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    ws.column_dimensions['B'].width = 50
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=questions_export.xlsx"}
    )


# ---- 分类管理接口 ----

@router.get("/categories")
def list_categories(db: Session = Depends(get_db), _=Depends(get_current_user)):
    """获取所有分类及其题目数量"""
    return crud.get_all_categories(db)


@router.post("/categories")
def create_category_endpoint(data: dict, db: Session = Depends(get_db), _=Depends(require_admin)):
    """创建新分类"""
    category_name = data.get("name", "").strip()
    if not category_name:
        raise HTTPException(status_code=400, detail="分类名称不能为空")
    success = crud.create_category(db, category_name)
    if not success:
        raise HTTPException(status_code=400, detail="分类名称已存在")
    return {"ok": True}


@router.put("/categories/rename")
def rename_category_endpoint(data: dict, db: Session = Depends(get_db), _=Depends(require_admin)):
    """重命名分类"""
    old_name = data.get("old_name", "").strip()
    new_name = data.get("new_name", "").strip()
    if not old_name or not new_name:
        raise HTTPException(status_code=400, detail="分类名称不能为空")
    if old_name == new_name:
        return {"ok": True}
    # 检查新名称是否已存在
    existing = crud.get_all_categories(db)
    if any(c["name"] == new_name for c in existing):
        raise HTTPException(status_code=400, detail="分类名称已存在")
    crud.rename_category(db, old_name, new_name)
    return {"ok": True}


@router.delete("/categories/{category_name:path}")
def delete_category_endpoint(category_name: str, db: Session = Depends(get_db), _=Depends(require_admin)):
    """删除分类（仅当无题目时）"""
    # URL decode category_name
    from urllib.parse import unquote
    category_name = unquote(category_name)
    categories = crud.get_all_categories(db)
    category = next((c for c in categories if c["name"] == category_name), None)
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")
    if category["count"] > 0:
        raise HTTPException(status_code=400, detail="该分类下还有题目，无法删除")
    # 由于没有独立的分类表，且该分类下无题目，直接返回成功
    # 实际删除逻辑：将所有该分类的题目（应该为0个）的 category 设为空
    crud.rename_category(db, category_name, "")
    return {"ok": True}


@router.post("/random")
def random_questions(data: dict, db: Session = Depends(get_db), _=Depends(require_admin)):
    """随机抽取指定数量的题目
    请求体: {"single": 5, "multiple": 3, "judgment": 2}
    返回: {"question_ids": [1, 2, 3, ...]}
    """
    import random
    from sqlalchemy import func

    single_count = data.get("single", 0)
    multiple_count = data.get("multiple", 0)
    judgment_count = data.get("judgment", 0)

    result_ids = []

    # 获取单选题
    if single_count > 0:
        single_ids = db.query(models.Question.id).filter(
            models.Question.type == "single"
        ).all()
        single_ids = [q.id for q in single_ids]
        if len(single_ids) < single_count:
            raise HTTPException(status_code=400, detail=f"单选题数量不足，题库中只有 {len(single_ids)} 道单选题")
        result_ids.extend(random.sample(single_ids, single_count))

    # 获取多选题
    if multiple_count > 0:
        multiple_ids = db.query(models.Question.id).filter(
            models.Question.type == "multiple"
        ).all()
        multiple_ids = [q.id for q in multiple_ids]
        if len(multiple_ids) < multiple_count:
            raise HTTPException(status_code=400, detail=f"多选题数量不足，题库中只有 {len(multiple_ids)} 道多选题")
        result_ids.extend(random.sample(multiple_ids, multiple_count))

    # 获取判断题
    if judgment_count > 0:
        judgment_ids = db.query(models.Question.id).filter(
            models.Question.type == "judgment"
        ).all()
        judgment_ids = [q.id for q in judgment_ids]
        if len(judgment_ids) < judgment_count:
            raise HTTPException(status_code=400, detail=f"判断题数量不足，题库中只有 {len(judgment_ids)} 道判断题")
        result_ids.extend(random.sample(judgment_ids, judgment_count))

    # 查询选中题目的详细信息（用于前端显示）
    selected_questions = db.query(models.Question.id, models.Question.type).filter(
        models.Question.id.in_(result_ids)
    ).all()

    return {
        "question_ids": result_ids,
        "questions": [{"id": q.id, "type": q.type} for q in selected_questions]
    }
